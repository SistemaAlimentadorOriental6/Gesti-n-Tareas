from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from admintaskG.models import Comite, Tareas, Subtareas, SubtareasAdicionales
from django.utils.timezone import now
from django.http import JsonResponse
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.core.mail import EmailMultiAlternatives, get_connection
from email.mime.image import MIMEImage
from io import BytesIO
from django.core.cache import cache
import threading
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side 
import os
from admintaskG import settings
from datetime import datetime
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from django.db.models import Q





def login(request):
    context = {}
    if request.method == 'POST':
        username = request.POST['usuario']
        password = request.POST['contrasena']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            auth_login(request, user)
            context['login_success'] = True  
        else:
            context['login_error'] = True  
    return render(request, 'Login/login.html', context)





def logout_view(request):
    logout(request)
    return redirect('login')





@login_required
def agregarComite(request):
    if request.method == "POST":
        descripcion_comite = request.POST.get("descripcion_comite")
        if descripcion_comite:
            Comite.objects.create(descripcion_comite=descripcion_comite, creador=request.user)
    return redirect('listadoComites')




@login_required
def listadoComites(request):
    comites = Comite.objects.all() 
    return render(request, "Comites/listadoComites.html", {"comites": comites})




def tareas_por_comite(request, comite_id):
    comite = get_object_or_404(Comite, id=comite_id)
    responsable_filtro = request.GET.get('responsable', '').strip()
    
    tareas = Tareas.objects.filter(comite_id=comite_id).order_by('-fecha_inicio')
    if responsable_filtro:
        tareas = tareas.filter(responsable__icontains=responsable_filtro)
    
    return render(request, "partials/listadoTareasParcial.html", {
        "comite": comite, 
        "tareas": tareas,
        "responsable_filtro": responsable_filtro
    })


def subtareas_por_tarea(request, tarea_id):
    tarea = get_object_or_404(Tareas, id=tarea_id)
    responsable_filtro = request.GET.get('responsable', '').strip()
    
    subtareas = Subtareas.objects.filter(tarea_id=tarea_id).order_by('-fecha_inicio')
    if responsable_filtro:
        subtareas = subtareas.filter(responsable__icontains=responsable_filtro)
    
    return render(request, "partials/listadoSubtareasParcial.html", {
        "subtareas": subtareas, 
        "tarea": tarea,
        "responsable_filtro": responsable_filtro
    })


def subtareasadicionales_por_subtarea(request, subtarea_id):
    subtarea = get_object_or_404(Subtareas, id=subtarea_id)
    responsable_filtro = request.GET.get('responsable', '').strip()
    
    subtareasadicionales = SubtareasAdicionales.objects.filter(subtarea_id=subtarea_id).order_by('-fecha_inicio')
    if responsable_filtro:
        subtareasadicionales = subtareasadicionales.filter(responsable__icontains=responsable_filtro)
    
    return render(request, "partials/listadoSubtareasAdicionalesParcial.html", {
        "subtareasadicionales": subtareasadicionales, 
        "subtarea": subtarea,
        "responsable_filtro": responsable_filtro
    })





@login_required
def listadoTareas(request, comite_id):
    comite = get_object_or_404(Comite, id=comite_id)
    tareas = Tareas.objects.filter(comite=comite).order_by("-id")
     
    return render(request, "Tareas/listadoTareas.html", {
        "tareas":tareas,
        "comite":comite
    })




@login_required
def agregarTarea(request):
    if request.method == "POST":
        comite_id = request.POST.get("comite_id")
        fecha_inicio = request.POST.get("fecha_inicio_tarea")
        fecha_cierre = request.POST.get("fecha_cierre") or None
        descripcion_tarea = request.POST.get("descripcion_tarea")
        responsable = request.POST.get("responsable")
        correo_responsable = request.POST.get("correo_responsable")

        try:
            comite = get_object_or_404(Comite, id=comite_id)

            Tareas.objects.create(
                comite=comite,
                fecha_inicio=fecha_inicio,
                fecha_cierre=fecha_cierre,
                descripcion_tarea=descripcion_tarea,
                responsable=responsable,
                correo_responsable=correo_responsable
            )

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": True})
            else:
                messages.success(request, "Tarea registrada correctamente.")
                return redirect('listadoTareas', comite_id=comite_id)

        except Exception as e:
            print(f"Error al registrar tarea: {e}")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({"success": False}, status=400)
            else:
                messages.error(request, "Hubo un error al registrar la tarea.")

    return redirect('listadoTareas', comite_id=comite_id)





@login_required
def listadoSubtareas(request, tarea_id):
    tarea = get_object_or_404(Tareas, id=tarea_id)
    subtareas = Subtareas.objects.filter(tarea=tarea).order_by("-id")
     
    return render(request, "Subtareas/listadoSubtareas.html", {
        "subtareas":subtareas,
        "tarea":tarea
    })
    
    
    

@login_required
def agregarSubtarea(request):
    if request.method == "POST":
        tarea_id = request.POST.get("tarea_id")
        fecha_inicio = request.POST.get("fecha_inicio_sub")
        fecha_cierre_raw = request.POST.get("fecha_cierre")
        descripcion_subtarea = request.POST.get("descripcion_subtarea")
        responsable = request.POST.get("responsable")
        correo_responsable = request.POST.get("correo_responsable")
        
        fecha_cierre = fecha_cierre_raw if fecha_cierre_raw else None

        
        try:
            tarea = get_object_or_404(Tareas, id=tarea_id)

            Subtareas.objects.create(
                tarea=tarea,
                fecha_inicio=fecha_inicio,
                fecha_cierre=fecha_cierre,
                descripcion_subtarea=descripcion_subtarea,
                responsable=responsable,
                correo_responsable=correo_responsable,
            )
            
            messages.success(request, "Subtarea registrada correctamente.")
            return redirect('listadoSubtareas')

        except Exception as e:
            print(f"Error al registrar subtarea: {e}")
            messages.error(request, "Hubo un error al registrar la subtarea.")
    
    return redirect('listadoSubtareas', tarea_id=tarea.id)




@login_required
def agregarSubtareaAdicional(request):
    if request.method == "POST":
        subtarea_id = request.POST.get("subtarea_id")
        fecha_inicio = request.POST.get("fecha_inicio_subad")
        fecha_cierre_raw = request.POST.get("fecha_cierre")
        descripcion_subtarea_adicional = request.POST.get("descripcion_subtarea_adicional")
        responsable = request.POST.get("responsable")
        correo_responsable = request.POST.get("correo_responsable")
        
        fecha_cierre = fecha_cierre_raw if fecha_cierre_raw else None
        
        try:
            subtarea = get_object_or_404(Subtareas, id=subtarea_id)
            
            SubtareasAdicionales.objects.create(
                subtarea=subtarea,
                fecha_inicio=fecha_inicio,
                fecha_cierre=fecha_cierre,
                descripcion_subtarea_adicional=descripcion_subtarea_adicional,
                responsable=responsable,
                correo_responsable=correo_responsable
            )
            
            messages.success(request, "Subtarea adicional registrada correctamente.")
            return render('listadoSubtareasAdicionales')
            
        except Exception as e:
            print(f"Error al registrar subtarea adicional: {e}")
            messages.error(request, "Hubo un error al registrar la subtarea adicional.")
            
    return redirect('listadoSubtareasAdicionales', subtarea_id=subtarea.id)




@login_required
@login_required
def listadoSubtareasAdicionales(request, subtarea_id):
    subtarea = get_object_or_404(Subtareas, id=subtarea_id)
    subtareas_adicionales = SubtareasAdicionales.objects.filter(subtarea=subtarea).order_by("-id")
     
    return render(request, "SubtareasAdicionales/listadoSubtareasAdicionales.html", {
        "subtareas_adicionales":subtareas_adicionales,
        "subtarea":subtarea
    })
    
    
     
    
@login_required
def estadoTarea(request, tarea_id):
    if request.method == "POST":
        try:
            tarea = Tareas.objects.get(id=tarea_id)
            tarea.estado = 0 if tarea.estado == 1 else 1

            if tarea.estado == 1:
                tarea.fecha_cierre = now().date()
            else:
                tarea.fecha_cierre = None  

            tarea.save()

            return JsonResponse({
                "nuevo_estado": tarea.estado,
                "nueva_fecha": tarea.fecha_cierre.strftime('%d/%m/%Y') if tarea.fecha_cierre else ""
            })
        except Tareas.DoesNotExist:
            return JsonResponse({"error": "Tarea no encontrada"}, status=404)
    return JsonResponse({"error": "Método no permitido"}, status=400)




@login_required
def estadoSubtarea(request, subtarea_id):
    if request.method == "POST":
        try:
            subtarea = Subtareas.objects.get(id=subtarea_id) 
            subtarea.estado = 0 if subtarea.estado == 1 else 1
            
            if subtarea.estado == 1:
                subtarea.fecha_cierre = now().date()
            else:
                subtarea.fecha_cierre = None 
            
            subtarea.save()
            return JsonResponse({
                "nuevo_estado": subtarea.estado,
                "nueva_fecha": subtarea.fecha_cierre.strftime('%d/%m/%Y') if subtarea.fecha_cierre else ""

            })
        except Subtareas.DoesNotExist:
            return JsonResponse({"error": "Subtarea no encontrada"}, status=404)
    return JsonResponse({"error": "Método no permitido"}, status=400)





@login_required
def estadoSubtareaAdicional(request, subtareaadicional_id):
    if request.method == "POST":
        try:
            subtarea_adicional = SubtareasAdicionales.objects.get(id=subtareaadicional_id) 
            subtarea_adicional.estado = 0 if subtarea_adicional.estado == 1 else 1
            
            if subtarea_adicional.estado == 1:
                subtarea_adicional.fecha_cierre = now().date()
            else:
                subtarea_adicional.fecha_cierre = None 
            
            subtarea_adicional.save()
            return JsonResponse({
                "nuevo_estado": subtarea_adicional.estado,
                "nueva_fecha": subtarea_adicional.fecha_cierre.strftime('%d/%m/%Y') if subtarea_adicional.fecha_cierre else ""

            })
        except SubtareasAdicionales.DoesNotExist:
            return JsonResponse({"error": "Subtarea adicional no encontrada"}, status=404)
    return JsonResponse({"error": "Método no permitido"}, status=400)




def enviar_pdf_individualizado(request, comite_id):
    """Vista principal que inicia el proceso de envío en segundo plano"""
    if request.method != 'POST':
        return JsonResponse({"error": "Método no permitido"}, status=405)
    
    print("🚀 Iniciando envío de correos para comité", comite_id)
    
    # Asegurar que tenemos una sesión
    if not request.session.session_key:
        request.session.save()
    session_key = request.session.session_key
    
    # Verificar si ya hay un proceso en ejecución
    cache_key = f"pdf_progreso_envio_{session_key}"
    estado_actual = cache.get(cache_key, {})
    
    if estado_actual.get("iniciado") and not estado_actual.get("completado"):
        return JsonResponse({
            "status": "ya_iniciado",
            "message": "Ya hay un proceso de envío en curso"
        })
    
    # Inicializar el progreso en caché
    cache.set(cache_key, {
        "progreso": 0,
        "total": 0,
        "completado": False,
        "resultados": [],
        "iniciado": True,
        "estado_actual": "Iniciando proceso..."
    }, timeout=600)
    
    # Ejecutar el envío en un hilo separado
    thread = threading.Thread(
        target=procesar_envio_correos, 
        args=(comite_id, session_key),
        daemon=True
    )
    thread.start()
    
    return JsonResponse({
        "status": "iniciado",
        "message": "Proceso de envío iniciado correctamente"
    })


def procesar_envio_correos(comite_id, session_key):
    """Función optimizada que procesa el envío de correos en segundo plano"""
    connection = None
    start_time = time.time()
    
    try:
        cache_key = f"pdf_progreso_envio_{session_key}"
        
        # Validar que el comité existe
        try:
            comite = Comite.objects.select_related().get(id=comite_id)
        except Comite.DoesNotExist:
            cache.set(cache_key, {
                "progreso": 0,
                "total": 0,
                "completado": True,
                "error": "Comité no encontrado",
                "resultados": []
            }, timeout=600)
            return

        # ---- Pre-cargar datos y recursos ----
        cache.set(cache_key, {
            "progreso": 2,
            "total": 0,
            "completado": False,
            "resultados": [],
            "estado_actual": "Cargando recursos..."
        }, timeout=600)

        # Cargar imagen de firma una sola vez
        firma_img = cargar_imagen_firma()
        
        # Preparar estilos de Excel una sola vez
        estilos_excel = preparar_estilos_excel()
        
        cache.set(cache_key, {
            "progreso": 5,
            "total": 0,
            "completado": False,
            "resultados": [],
            "estado_actual": "Recopilando datos de tareas..."
        }, timeout=600)

        # ---- Recopilar datos de manera optimizada ----
        datos_por_correo = recopilar_datos_tareas(comite)
        
        # Filtrar y validar correos
        datos_por_correo = {k: v for k, v in datos_por_correo.items() 
                           if k and '@' in k and len(k) > 5}
        
        total = len(datos_por_correo)

        if total == 0:
            cache.set(cache_key, {
                "progreso": 100,
                "total": 0,
                "completado": True,
                "resultados": [],
                "estado_actual": "No se encontraron correos válidos para enviar.",
                "error": "No hay correos válidos para procesar"
            }, timeout=600)
            return

        cache.set(cache_key, {
            "progreso": 10,
            "total": total,
            "completado": False,
            "resultados": [],
            "estado_actual": f"Preparando envío a {total} personas..."
        }, timeout=600)

        # ---- Envío paralelo optimizado ----
        resultados = enviar_correos_paralelo(
            datos_por_correo, comite, firma_img, estilos_excel, cache_key, session_key
        )

        # Estado final
        tiempo_total = round(time.time() - start_time, 2)
        cache.set(cache_key, {
            "progreso": 100,
            "total": total,
            "completado": True,
            "resultados": resultados,
            "estado_actual": f"¡Proceso completado! ({tiempo_total}s - {len(resultados)} correos)",
            "tiempo_ejecucion": tiempo_total
        }, timeout=600)

        print(f"✅ Proceso completado en {tiempo_total}s - {len(resultados)} correos enviados")

    except Exception as e:
        print(f"❌ Error en proceso principal: {e}")
        cache.set(cache_key, {
            "progreso": 0,
            "total": 0,
            "completado": True,
            "error": f"Error en el proceso: {str(e)}",
            "resultados": [],
            "estado_actual": "Error en el proceso"
        }, timeout=600)


def cargar_imagen_firma():
    """Carga la imagen de firma una sola vez"""
    firma_path = os.path.join(settings.BASE_DIR, "admintaskG", "Public", "Img", "logocorreo.png")
    if os.path.exists(firma_path):
        try:
            with open(firma_path, 'rb') as f:
                return f.read()
        except Exception as e:
            print(f"Error cargando imagen de firma: {e}")
    return None


def preparar_estilos_excel():
    """Prepara los estilos de Excel una sola vez"""
    return {
        'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')),
        'fill_header': PatternFill(start_color="6BA43A", end_color="6BA43A", fill_type="solid"),
        'fill_estado_pend': PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        'fill_estado_ok': PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        'font_header': Font(bold=True, color="FFFFFF"),
        'align_center': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'align_left': Alignment(horizontal="left", vertical="top", wrap_text=True),
        'column_widths': [5, 10, 15, 25, 30, 30, 15, 15, 30, 30, 15, 15, 30, 30, 15, 15],
        'headers': [
            "N°", "ID Tarea", "Estado", "Área", "Tarea", "Responsable Tarea", 
            "Fecha Inicio Tarea", "Fecha Cierre Tarea", "Subtarea", "Responsable Subtarea", 
            "Fecha Inicio Subtarea", "Fecha Cierre Subtarea", "Subtarea Adicional", 
            "Responsable Subtarea Adicional", "Fecha Inicio Subtarea Adicional", "Fecha Cierre Subtarea Adicional"
        ]
    }


def formatear_fecha(fecha):
    """Formatea fecha para mostrar en Excel de manera eficiente"""
    if not fecha:
        return ""
    try:
        if isinstance(fecha, str):
            return fecha
        return fecha.strftime("%Y-%m-%d") if fecha else ""
    except:
        return str(fecha) if fecha else ""


def recopilar_datos_tareas(comite):
    """Recopila todos los datos de tareas de manera optimizada con una sola consulta por tipo"""
    datos_por_correo = {}
    
    # ===== TAREAS - Consulta optimizada =====
    tareas = Tareas.objects.filter(
        comite=comite, estado=0
    ).select_related().only(
        'id', 'descripcion_tarea', 'estado', 'fecha_inicio', 
        'fecha_cierre', 'responsable', 'correo_responsable'
    )
    
    for tarea in tareas:
        correo = tarea.correo_responsable.strip() if tarea.correo_responsable else ""
        if correo:
            agregar_datos_correo(datos_por_correo, correo, tarea.responsable)
            
            if tarea.id not in datos_por_correo[correo]["tareas"]:
                datos_por_correo[correo]["tareas"][tarea.id] = {
                    "id": tarea.id,
                    "descripcion": tarea.descripcion_tarea or "",
                    "estado": tarea.estado,
                    "fecha_inicio": formatear_fecha(getattr(tarea, "fecha_inicio", "")),
                    "fecha_cierre": formatear_fecha(getattr(tarea, "fecha_cierre", "")),
                    "responsable_tarea": tarea.responsable or "",
                    "subtareas": {}
                }

    # ===== SUBTAREAS - Consulta optimizada =====
    subtareas = Subtareas.objects.filter(
        tarea__comite=comite, estado=0
    ).select_related('tarea').only(
        'id', 'descripcion_subtarea', 'estado', 'fecha_inicio', 'fecha_cierre',
        'responsable', 'correo_responsable', 'tarea__id', 'tarea__descripcion_tarea',
        'tarea__estado', 'tarea__fecha_inicio', 
        'tarea__fecha_cierre', 'tarea__responsable'
    )
    
    for subtarea in subtareas:
        correo = subtarea.correo_responsable.strip() if subtarea.correo_responsable else ""
        if correo:
            agregar_datos_correo(datos_por_correo, correo, subtarea.responsable)
            tarea_id = subtarea.tarea.id
            
            # Asegurar tarea padre
            if tarea_id not in datos_por_correo[correo]["tareas"]:
                datos_por_correo[correo]["tareas"][tarea_id] = crear_datos_tarea_padre(subtarea.tarea)
            
            # Agregar subtarea
            datos_por_correo[correo]["tareas"][tarea_id]["subtareas"][subtarea.id] = {
                "descripcion": subtarea.descripcion_subtarea or "",
                "estado": subtarea.estado,
                "fecha_inicio": formatear_fecha(getattr(subtarea, "fecha_inicio", "")),
                "fecha_cierre": formatear_fecha(getattr(subtarea, "fecha_cierre", "")),
                "responsable_subtarea": subtarea.responsable or "",
                "subadicionales": []
            }

    # ===== SUBTAREAS ADICIONALES - Consulta optimizada =====
    adicionales = SubtareasAdicionales.objects.filter(
        subtarea__tarea__comite=comite, estado=0
    ).select_related('subtarea__tarea').only(
        'descripcion_subtarea_adicional', 'estado', 'fecha_inicio', 'fecha_cierre',
        'responsable', 'correo_responsable', 'subtarea__id', 'subtarea__descripcion_subtarea',
        'subtarea__estado', 'subtarea__fecha_inicio', 'subtarea__fecha_cierre',
        'subtarea__responsable', 'subtarea__tarea__id', 'subtarea__tarea__descripcion_tarea',
        'subtarea__tarea__estado', 'subtarea__tarea__fecha_inicio', 'subtarea__tarea__fecha_cierre', 'subtarea__tarea__responsable'
    )
    
    for adicional in adicionales:
        correo = adicional.correo_responsable.strip() if adicional.correo_responsable else ""
        if correo:
            agregar_datos_correo(datos_por_correo, correo, adicional.responsable)
            tarea_id = adicional.subtarea.tarea.id
            subtarea_id = adicional.subtarea.id
            
            # Asegurar tarea padre
            if tarea_id not in datos_por_correo[correo]["tareas"]:
                datos_por_correo[correo]["tareas"][tarea_id] = crear_datos_tarea_padre(adicional.subtarea.tarea)
            
            # Asegurar subtarea padre
            if subtarea_id not in datos_por_correo[correo]["tareas"][tarea_id]["subtareas"]:
                datos_por_correo[correo]["tareas"][tarea_id]["subtareas"][subtarea_id] = {
                    "descripcion": adicional.subtarea.descripcion_subtarea or "",
                    "estado": adicional.subtarea.estado,
                    "fecha_inicio": formatear_fecha(getattr(adicional.subtarea, "fecha_inicio", "")),
                    "fecha_cierre": formatear_fecha(getattr(adicional.subtarea, "fecha_cierre", "")),
                    "responsable_subtarea": adicional.subtarea.responsable or "",
                    "subadicionales": []
                }
            
            # Agregar subtarea adicional
            datos_por_correo[correo]["tareas"][tarea_id]["subtareas"][subtarea_id]["subadicionales"].append({
                "descripcion": adicional.descripcion_subtarea_adicional or "",
                "estado": adicional.estado,
                "fecha_inicio": formatear_fecha(getattr(adicional, "fecha_inicio", "")),
                "fecha_cierre": formatear_fecha(getattr(adicional, "fecha_cierre", "")),
                "responsable": adicional.responsable or ""
            })
    
    return datos_por_correo


def agregar_datos_correo(datos_por_correo, correo, responsable):
    """Helper para agregar datos de correo sin duplicación"""
    if correo not in datos_por_correo:
        datos_por_correo[correo] = {
            "responsable": responsable or "Sin responsable",
            "tareas": {}
        }


def crear_datos_tarea_padre(tarea):
    """Helper para crear datos de tarea padre"""
    return {
        "id": tarea.id,
        "descripcion": tarea.descripcion_tarea or "",
        "estado": tarea.estado,
        "fecha_inicio": formatear_fecha(getattr(tarea, "fecha_inicio", "")),
        "fecha_cierre": formatear_fecha(getattr(tarea, "fecha_cierre", "")),
        "responsable_tarea": tarea.responsable or "",
        "subtareas": {}
    }


def enviar_correos_paralelo(datos_por_correo, comite, firma_img, estilos_excel, cache_key, session_key):
    """Envía correos de forma paralela para máxima velocidad"""
    resultados = []
    total = len(datos_por_correo)
    
    # Configurar pool de threads (máximo 5 conexiones simultáneas para evitar sobrecarga del servidor SMTP)
    max_workers = min(5, total)
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Crear tareas de envío
        future_to_correo = {}
        
        for i, (correo, info) in enumerate(datos_por_correo.items()):
            future = executor.submit(
                enviar_correo_individual, 
                correo, info, comite, firma_img, estilos_excel
            )
            future_to_correo[future] = (correo, info, i + 1)
        
        # Procesar resultados conforme se completan
        completados = 0
        for future in as_completed(future_to_correo):
            correo, info, numero = future_to_correo[future]
            completados += 1
            
            try:
                resultado = future.result(timeout=30)  # Timeout de 30s por correo
                resultado["numero"] = numero
                resultados.append(resultado)
                
                # Actualizar progreso
                progreso_actual = min(int(10 + ((completados / total) * 90)), 100)
                cache.set(cache_key, {
                    "progreso": progreso_actual,
                    "total": total,
                    "completado": False,
                    "resultados": sorted(resultados, key=lambda x: x.get("numero", 0)),
                    "estado_actual": f"Enviado {completados}/{total} - {info.get('responsable', 'Usuario')}"
                }, timeout=600)
                
            except Exception as e:
                print(f"❌ Error enviando a {correo}: {e}")
                resultado = {
                    "correo": correo,
                    "estado": f"❌ Error: {str(e)[:100]}...",
                    "responsable": info.get("responsable", "Desconocido"),
                    "timestamp": timezone.now().strftime("%H:%M:%S"),
                    "numero": numero
                }
                resultados.append(resultado)
    
    return sorted(resultados, key=lambda x: x.get("numero", 0))


def enviar_correo_individual(correo, info, comite, firma_img, estilos_excel):
    """Envía un correo individual de manera optimizada"""
    start_time = time.time()
    
    try:
        # Validar correo
        validate_email(correo)
        
        # Generar Excel en memoria
        excel_content = generar_excel_optimizado(info, comite, estilos_excel)
        
        # Crear y enviar email
        enviar_email_con_adjunto(correo, comite, excel_content, firma_img)
        
        # Calcular estadísticas
        stats = calcular_estadisticas(info)
        tiempo_envio = round(time.time() - start_time, 2)
        
        return {
            "correo": correo,
            "estado": f"✅ Enviado ({tiempo_envio}s)",
            "responsable": info["responsable"],
            "tareas": stats["tareas"],
            "subtareas": stats["subtareas"],
            "subadicionales": stats["subadicionales"],
            "timestamp": timezone.now().strftime("%H:%M:%S"),
            "tiempo_envio": tiempo_envio
        }
        
    except ValidationError:
        return {
            "correo": correo,
            "estado": "❌ Correo inválido",
            "responsable": info.get("responsable", "Desconocido"),
            "timestamp": timezone.now().strftime("%H:%M:%S")
        }
    except Exception as e:
        return {
            "correo": correo,
            "estado": f"❌ Error: {str(e)[:80]}...",
            "responsable": info.get("responsable", "Desconocido"),
            "timestamp": timezone.now().strftime("%H:%M:%S")
        }


def generar_excel_optimizado(info, comite, estilos):
    """Genera Excel de manera optimizada"""
    excel_file = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas Comité"
    
    # Headers
    ws.append(estilos['headers'])
    for col in ws[1]:
        col.fill = estilos['fill_header']
        col.font = estilos['font_header']
        col.alignment = estilos['align_center']
        col.border = estilos['border']
    
    # Datos
    contador_filas = 1
    for tarea in info["tareas"].values():
        responsable_tarea = tarea.get("responsable_tarea", "")
        
        if not tarea["subtareas"]:
            # Tarea sin subtareas
            ws.append([
                contador_filas, tarea.get("id", ""),
                "Pendiente" if tarea["estado"] == 0 else "Completado",
                comite.descripcion_comite or "", tarea["descripcion"],
                responsable_tarea, tarea.get("fecha_inicio", ""),
                tarea.get("fecha_cierre", ""), "", "", "", "", "", "", "", ""
            ])
            contador_filas += 1
        else:
            for subtarea in tarea["subtareas"].values():
                responsable_subtarea = subtarea.get("responsable_subtarea", "")
                
                if not subtarea.get("subadicionales", []):
                    # Subtarea sin adicionales
                    ws.append([
                        contador_filas, tarea.get("id", ""),
                        "Pendiente" if subtarea["estado"] == 0 else "Completado",
                        comite.descripcion_comite or "", tarea["descripcion"],
                        responsable_tarea, tarea.get("fecha_inicio", ""),
                        tarea.get("fecha_cierre", ""), subtarea["descripcion"],
                        responsable_subtarea, subtarea.get("fecha_inicio", ""),
                        subtarea.get("fecha_cierre", ""), "", "", "", ""
                    ])
                    contador_filas += 1
                else:
                    for adicional in subtarea["subadicionales"]:
                        ws.append([
                            contador_filas, tarea.get("id", ""),
                            "Pendiente" if adicional["estado"] == 0 else "Completado",
                            comite.descripcion_comite or "", tarea["descripcion"],
                            responsable_tarea, tarea.get("fecha_inicio", ""),
                            tarea.get("fecha_cierre", ""), subtarea["descripcion"],
                            responsable_subtarea, subtarea.get("fecha_inicio", ""),
                            subtarea.get("fecha_cierre", ""), adicional["descripcion"],
                            adicional["responsable"], adicional.get("fecha_inicio", ""),
                            adicional.get("fecha_cierre", "")
                        ])
                        contador_filas += 1
    
    # Aplicar formato de manera eficiente
    aplicar_formato_excel(ws, estilos)
    
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.read()


def aplicar_formato_excel(ws, estilos):
    """Aplica formato al Excel de manera optimizada"""
    # Formato de filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = estilos['border']
            if cell.column in [1, 2, 3, 7, 8, 11, 12, 15, 16]:
                cell.alignment = estilos['align_center']
            else:
                cell.alignment = estilos['align_left']
        
        # Color según estado
        if len(row) > 2:
            estado_cell = row[2]
            if estado_cell.value == "Pendiente":
                estado_cell.fill = estilos['fill_estado_pend']
            elif estado_cell.value == "Completado":
                estado_cell.fill = estilos['fill_estado_ok']
    
    # Ajustar ancho de columnas
    for idx, width in enumerate(estilos['column_widths'], 1):
        if idx <= len(estilos['headers']):
            ws.column_dimensions[get_column_letter(idx)].width = width


def enviar_email_con_adjunto(correo, comite, excel_content, firma_img):
    """Envía email con adjunto de manera optimizada"""
    # Crear conexión individual (más rápido para envío paralelo)
    connection = get_connection(fail_silently=False)
    
    try:
        connection.open()
        
        # Preparar datos del email
        nombre_comite = (comite.descripcion_comite or "Comite").replace(" ", "_").replace("/", "_")
        fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M")
        nombre_archivo = f"{nombre_comite}_{fecha_actual}.xlsx"
        
        subject = f"Listado de tareas - {comite.descripcion_comite or 'Comité'}"
        from_email = settings.DEFAULT_FROM_EMAIL
        to = [correo]
        
        html_content = """
        <div style="font-family: Arial, sans-serif; font-size: 14px; color: #333; line-height: 1.6;">
            <p style="margin-bottom: 15px;">Cordial saludo,</p>
            <p style="margin-bottom: 15px;">Comparto pendientes correspondientes a la reunión.</p>
            <p style="margin-bottom: 15px;">Muchas gracias.</p>
            <p style="margin-bottom: 15px;">Cordialmente,</p>
            <div style="padding: 15px; border-left: 4px solid #28a745; background-color: #f9f9f9; max-width: 500px;">
                <p style="margin: 0;">
                    <strong style="font-size: 16px;">Melisa Rojas</strong><br>
                    <span style="color: #28a745;">Analista de Mejora Continua</span><br>
                    <span style="color: #555;">350 495 88 27</span><br>
                    <span style="color: #555;">Cra. 56 A N° 62 – 50 - Medellín</span>
                </p>
            </div>
            <div style="margin-top: 15px;">
                <img src="cid:firma_banner" alt="Firma" style="max-width: 600px; border-radius: 4px; display: block;">
            </div>
        </div>
        """
        
        # Crear email
        email = EmailMultiAlternatives(subject, "", from_email, to, connection=connection)
        email.attach_alternative(html_content, "text/html")
        email.attach(nombre_archivo, excel_content,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # Adjuntar imagen si existe
        if firma_img:
            try:
                img = MIMEImage(firma_img)
                img.add_header('Content-ID', '<firma_banner>')
                img.add_header('Content-Disposition', 'inline')
                email.attach(img)
            except Exception as e:
                print(f"Error adjuntando imagen: {e}")
        
        # Enviar
        email.send(fail_silently=False)
        
    finally:
        connection.close()


def calcular_estadisticas(info):
    """Calcula estadísticas de manera optimizada"""
    total_tareas = len(info["tareas"])
    total_subtareas = sum(len(t["subtareas"]) for t in info["tareas"].values())
    total_subadicionales = sum(len(s.get("subadicionales", []))
                              for t in info["tareas"].values()
                              for s in t["subtareas"].values())
    
    return {
        "tareas": total_tareas,
        "subtareas": total_subtareas,
        "subadicionales": total_subadicionales
    }


def obtener_estado_envio(request):
    """Vista optimizada para obtener el estado actual del envío"""
    session_key = request.session.session_key
    if not session_key:
        return JsonResponse({
            "progreso": 0,
            "total": 0,
            "completado": False,
            "resultados": [],
            "error": "Sesión no válida"
        })
    
    cache_key = f"pdf_progreso_envio_{session_key}"
    estado = cache.get(cache_key, {
        "progreso": 0,
        "total": 0,
        "completado": False,
        "resultados": [],
        "estado_actual": "Esperando..."
    })
    
    return JsonResponse(estado)








def filtrar_por_responsable(request, comite_id):
    comite = get_object_or_404(Comite, id=comite_id)
    responsable_filtro = request.GET.get('responsable', '').strip()
    
    if not responsable_filtro:
        return JsonResponse({'error': 'Debe especificar un responsable'})
    
    # Filtrar tareas
    tareas = Tareas.objects.filter(
        comite_id=comite_id,
        responsable__icontains=responsable_filtro
    )
    
    # Filtrar subtareas de todas las tareas del comité
    subtareas = Subtareas.objects.filter(
        tarea__comite_id=comite_id,
        responsable__icontains=responsable_filtro
    )
    
    # Filtrar subtareas adicionales de todas las subtareas del comité
    subtareas_adicionales = SubtareasAdicionales.objects.filter(
        subtarea__tarea__comite_id=comite_id,
        responsable__icontains=responsable_filtro
    )
    
    context = {
        'comite': comite,
        'tareas': tareas,
        'subtareas': subtareas,
        'subtareas_adicionales': subtareas_adicionales,
        'responsable_filtro': responsable_filtro,
        'es_filtrado': True
    }
    
    return render(request, "partials/filtradoCompleto.html", context)